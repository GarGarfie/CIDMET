import os
import re
import csv
import io
from typing import List, Dict, Tuple, Optional

import xlrd
import xlwt
import openpyxl

from parsers import ParsedRecord
from matcher import MatchResult
from utils import detect_encoding, read_file_text, detect_csv_dialect


# ---------------------------------------------------------------------------
# Subset Writers - Format Preserving
# ---------------------------------------------------------------------------

def write_wos_txt_subset(
    matched_indices: List[int],
    raw_blocks: List[str],
    header: str,
    footer: str,
    output_path: str,
    encoding: str = 'utf-8',
):
    with open(output_path, 'w', encoding=encoding, errors='replace') as f:
        f.write(header)
        for idx in matched_indices:
            if 0 <= idx < len(raw_blocks):
                block = raw_blocks[idx]
                if not block.startswith('\n'):
                    f.write('\n')
                f.write(block)
                if not block.endswith('\n'):
                    f.write('\n')
        f.write('\nEF')


def write_wos_xls_subset(
    matched_row_indices: List[int],
    source_workbook,
    headers: List[str],
    output_path: str,
):
    src_ws = source_workbook.sheet_by_index(0)

    try:
        wb_out = xlwt.Workbook(encoding='utf-8')
        ws_out = wb_out.add_sheet('Sheet1')
        for c, h in enumerate(headers):
            ws_out.write(0, c, h)
        out_row = 1
        for row_idx in matched_row_indices:
            if 1 <= row_idx < src_ws.nrows:
                for c in range(src_ws.ncols):
                    val = src_ws.cell_value(row_idx, c)
                    ws_out.write(out_row, c, val)
                out_row += 1
        wb_out.save(output_path)
    except Exception:
        if not output_path.endswith('.xlsx'):
            output_path = output_path.rsplit('.', 1)[0] + '.xlsx'
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.append(headers)
        for row_idx in matched_row_indices:
            if 1 <= row_idx < src_ws.nrows:
                row_data = [src_ws.cell_value(row_idx, c) for c in range(src_ws.ncols)]
                ws_out.append(row_data)
        wb_out.save(output_path)

    return output_path


def _safe_dialect(dialect):
    """Ensure dialect has valid escape/quoting settings to avoid csv.Error."""
    class SafeDialect(csv.Dialect):
        delimiter = getattr(dialect, 'delimiter', ',')
        quotechar = getattr(dialect, 'quotechar', '"')
        doublequote = True
        skipinitialspace = getattr(dialect, 'skipinitialspace', False)
        lineterminator = getattr(dialect, 'lineterminator', '\r\n')
        quoting = getattr(dialect, 'quoting', csv.QUOTE_MINIMAL)
    return SafeDialect()


def write_csv_subset(
    matched_row_indices: List[int],
    raw_data_rows: List[List[str]],
    headers: List[str],
    dialect,
    encoding: str,
    output_path: str,
):
    safe = _safe_dialect(dialect)
    with open(output_path, 'w', encoding=encoding, errors='replace', newline='') as f:
        writer = csv.writer(f, dialect=safe)
        writer.writerow(headers)
        for row_idx in matched_row_indices:
            actual_idx = row_idx - 1 if row_idx >= 1 else row_idx
            if 0 <= actual_idx < len(raw_data_rows):
                writer.writerow(raw_data_rows[actual_idx])


def write_scopus_txt_subset(
    matched_indices: List[int],
    raw_blocks: List[str],
    header_text: str,
    output_path: str,
    encoding: str = 'utf-8',
):
    with open(output_path, 'w', encoding=encoding, errors='replace') as f:
        f.write(header_text)
        if not header_text.endswith('\n'):
            f.write('\n')
        for i, idx in enumerate(matched_indices):
            if 0 <= idx < len(raw_blocks):
                block = raw_blocks[idx]
                f.write(block)
                if not block.endswith('\n'):
                    f.write('\n')
                f.write('\n')


def write_ei_txt_subset(
    matched_indices: List[int],
    raw_blocks: List[str],
    output_path: str,
    encoding: str = 'utf-8',
):
    with open(output_path, 'w', encoding=encoding, errors='replace') as f:
        renumber = 1
        for idx in matched_indices:
            if 0 <= idx < len(raw_blocks):
                block = raw_blocks[idx]
                block = re.sub(r'<RECORD\s+\d+>', f'<RECORD {renumber}>', block, count=1)
                f.write(block)
                if not block.endswith('\n'):
                    f.write('\n')
                f.write('\n')
                renumber += 1


# ---------------------------------------------------------------------------
# Merge Export
# ---------------------------------------------------------------------------

UNIFIED_FIELDS = [
    'title', 'authors', 'author_full', 'year', 'journal', 'volume', 'issue',
    'pages', 'doi', 'issn', 'isbn', 'abstract', 'author_keywords',
    'index_keywords', 'document_type', 'language', 'publisher', 'affiliations',
    'funding', 'references', 'source_db',
    # Extended fields
    'cited_by', 'correspondence', 'article_number', 'page_start', 'page_end',
    'abbreviated_source', 'eissn', 'open_access', 'conference_name',
    'conference_date', 'conference_location', 'conference_code', 'sponsors',
    'editors', 'author_ids', 'researcher_ids', 'publication_date', 'link',
    'accession_number', 'funding_text', 'publication_type', 'num_references',
    'page_count', 'research_areas', 'email', 'authors_with_affiliations',
    'coden', 'pubmed_id', 'classification_code', 'main_heading',
    'reprint_address', 'publisher_address', 'publisher_city',
    'journal_abbreviation', 'journal_iso', 'organization_enhanced',
    'publication_stage', 'wos_id', 'usage_count_180', 'usage_count_since2013',
]


def record_to_unified(rec: ParsedRecord) -> dict:
    u = {f: '' for f in UNIFIED_FIELDS}
    u['title'] = rec.title
    u['authors'] = rec.authors
    u['year'] = rec.year
    u['journal'] = rec.journal
    u['volume'] = rec.volume
    u['pages'] = rec.pages
    u['doi'] = rec.doi
    u['issn'] = rec.issn
    u['isbn'] = rec.isbn
    u['abstract'] = rec.abstract
    u['author_keywords'] = rec.keywords
    u['source_db'] = rec.source_db
    u['source_format'] = rec.source_format

    rf = rec.raw_fields

    # Store ALL raw fields for pass-through export
    u['_raw_fields'] = dict(rf)

    if rec.source_db == 'wos' and rec.source_format == 'txt':
        u['issue'] = rf.get('IS', '')
        u['document_type'] = rf.get('DT', '')
        u['language'] = rf.get('LA', '')
        u['publisher'] = rf.get('PU', '')
        u['affiliations'] = rf.get('C1', '')
        u['funding'] = rf.get('FU', '')
        u['references'] = rf.get('CR', '')
        u['index_keywords'] = rf.get('ID', '')
        u['author_full'] = rf.get('AF', '')
        u['authors_raw_short'] = rf.get('AU', '')
        # Extended fields
        u['cited_by'] = rf.get('TC', '')
        u['correspondence'] = rf.get('RP', '')
        u['article_number'] = rf.get('AR', '')
        u['eissn'] = rf.get('EI', '')
        u['open_access'] = rf.get('OA', '')
        u['publication_date'] = rf.get('PD', '')
        u['publication_type'] = rf.get('PT', '')
        u['num_references'] = rf.get('NR', '')
        u['page_count'] = rf.get('PG', '')
        u['research_areas'] = rf.get('SC', '')
        u['email'] = rf.get('EM', '')
        u['researcher_ids'] = rf.get('RI', '')
        u['accession_number'] = rf.get('UT', '')
        u['wos_id'] = rf.get('UT', '')
        u['funding_text'] = rf.get('FX', '')
        u['reprint_address'] = rf.get('RP', '')
        u['publisher_city'] = rf.get('PI', '')
        u['publisher_address'] = rf.get('PA', '')
        u['journal_abbreviation'] = rf.get('J9', '')
        u['journal_iso'] = rf.get('JI', '')
        u['organization_enhanced'] = rf.get('C3', '')
        u['usage_count_180'] = rf.get('U1', '')
        u['usage_count_since2013'] = rf.get('U2', '')
        bp = rf.get('BP', '')
        ep = rf.get('EP', '')
        if bp:
            u['page_start'] = bp
        if ep:
            u['page_end'] = ep
        u['abbreviated_source'] = rf.get('JI', '') or rf.get('J9', '')

    elif rec.source_db == 'wos' and rec.source_format == 'xls':
        u['issue'] = rf.get('Issue', '')
        u['document_type'] = rf.get('Document Type', '')
        u['language'] = rf.get('Language', '')
        u['publisher'] = rf.get('Publisher', '')
        u['affiliations'] = rf.get('Addresses', '')
        u['funding'] = rf.get('Funding Orgs', '')
        u['references'] = rf.get('Cited References', '')
        u['index_keywords'] = rf.get('Keywords Plus', '')
        u['author_full'] = rf.get('Author Full Names', '')
        u['authors_raw_short'] = rf.get('Authors', '')
        # Extended fields
        u['cited_by'] = rf.get('Times Cited, All Databases', rf.get('Times Cited, WoS Core', ''))
        u['article_number'] = rf.get('Article Number', '')
        u['eissn'] = rf.get('eISSN', '')
        u['open_access'] = rf.get('Open Access Designations', '')
        u['publication_date'] = rf.get('Publication Date', '')
        u['publication_type'] = rf.get('Publication Type', '')
        u['num_references'] = rf.get('Number of Cited References', '')
        u['page_count'] = rf.get('Page Count', '')
        u['research_areas'] = rf.get('Research Areas', rf.get('Web of Science Categories', ''))
        u['email'] = rf.get('Email Addresses', '')
        u['researcher_ids'] = rf.get('ResearcherID Number', '')
        u['accession_number'] = rf.get('UT (Unique WOS ID)', '')
        u['wos_id'] = rf.get('UT (Unique WOS ID)', '')
        u['funding_text'] = rf.get('Funding Text', '')
        u['correspondence'] = rf.get('Reprint Addresses', '')
        u['reprint_address'] = rf.get('Reprint Addresses', '')
        u['publisher_address'] = rf.get('Publisher Address', '')
        u['publisher_city'] = rf.get('Publisher City', '')
        u['journal_abbreviation'] = rf.get('Journal Abbreviation', '')
        u['journal_iso'] = rf.get('Journal ISO Abbreviation', '')
        u['page_start'] = rf.get('Start Page', '')
        u['page_end'] = rf.get('End Page', '')
        u['abbreviated_source'] = rf.get('Journal ISO Abbreviation', rf.get('Journal Abbreviation', ''))
        u['editors'] = rf.get('Book Editors', '')

    elif rec.source_db == 'scopus' and rec.source_format == 'csv':
        u['issue'] = rf.get('Issue', rf.get('期', ''))
        u['document_type'] = rf.get('Document Type', rf.get('文献类型', ''))
        u['language'] = rf.get('Language of Original Document', rf.get('原始文献语言', ''))
        u['publisher'] = rf.get('Publisher', rf.get('出版商', ''))
        u['affiliations'] = rf.get('Affiliations', rf.get('归属机构', ''))
        u['funding'] = rf.get('Funding Details', rf.get('出资详情', ''))
        u['references'] = rf.get('References', rf.get('参考文献', ''))
        u['index_keywords'] = rf.get('Index Keywords', rf.get('索引关键字', ''))
        u['author_full'] = rf.get('Author full names', '')
        u['authors_raw_short'] = rec.authors
        # Extended fields
        u['cited_by'] = rf.get('Cited by', rf.get('引用次数', ''))
        u['correspondence'] = rf.get('Correspondence Address', rf.get('通讯地址', ''))
        u['article_number'] = rf.get('Art. No.', rf.get('文献编号', ''))
        u['page_start'] = rf.get('Page start', rf.get('起始页码', ''))
        u['page_end'] = rf.get('Page end', rf.get('结束页码', ''))
        u['abbreviated_source'] = rf.get('Abbreviated Source Title', rf.get('来源出版物名称缩写', ''))
        u['eissn'] = rf.get('eISSN', '')
        u['open_access'] = rf.get('Open Access', rf.get('开放获取', ''))
        u['conference_name'] = rf.get('Conference name', rf.get('会议名称', ''))
        u['conference_date'] = rf.get('Conference date', rf.get('会议日期', ''))
        u['conference_location'] = rf.get('Conference location', rf.get('会议地点', ''))
        u['conference_code'] = rf.get('Conference code', rf.get('会议代码', ''))
        u['sponsors'] = rf.get('Sponsors', rf.get('主办方', ''))
        u['editors'] = rf.get('Editors', rf.get('编辑', ''))
        u['author_ids'] = rf.get('Author(s) ID', rf.get('作者标识号', ''))
        u['link'] = rf.get('Link', rf.get('链接', ''))
        u['accession_number'] = rf.get('EID', '')
        u['funding_text'] = rf.get('Funding Texts', rf.get('出资文本', ''))
        u['publication_stage'] = rf.get('Publication Stage', rf.get('出版阶段', ''))
        u['authors_with_affiliations'] = rf.get('Authors with affiliations', rf.get('含归属机构的作者', ''))
        u['coden'] = rf.get('CODEN', '')
        u['pubmed_id'] = rf.get('PubMed ID', '')

    elif rec.source_db == 'scopus' and rec.source_format == 'txt':
        u['document_type'] = rf.get('document_type', '')
        u['language'] = rf.get('language', '')
        u['publisher'] = rf.get('publisher', '')
        u['affiliations'] = rf.get('affiliations', '')
        u['funding'] = rf.get('funding_details', '')
        u['references'] = rf.get('references', '')
        u['index_keywords'] = rf.get('index_keywords', '')
        u['author_full'] = rf.get('author_full', '')
        u['authors_raw_short'] = rf.get('author_short', '')
        # Extended fields
        u['correspondence'] = rf.get('correspondence', '')
        u['abbreviated_source'] = rf.get('abbreviated_source', '')
        u['open_access'] = rf.get('open_access', '')
        u['publication_stage'] = rf.get('publication_stage', '')
        u['funding_text'] = rf.get('funding_text', '')
        u['link'] = rf.get('link', '')
        # Extract EID from link URL for Scopus TXT
        link = rf.get('link', '')
        if link and not u['accession_number']:
            eid_match = re.search(r'eid=(2-s2\.0-\d+)', link)
            if eid_match:
                u['accession_number'] = eid_match.group(1)

    elif rec.source_db == 'ei' and rec.source_format == 'csv':
        u['issue'] = rf.get('Issue', '')
        u['document_type'] = rf.get('Document type', '')
        u['language'] = rf.get('Language', '')
        u['publisher'] = rf.get('Publisher/Repository', '')
        u['affiliations'] = rf.get('Author affiliation', '')
        u['funding'] = rf.get('Funding details', '')
        u['index_keywords'] = rf.get('Controlled/Subject terms', '')
        u['authors_raw_short'] = rec.authors
        # Extended fields
        u['accession_number'] = rf.get('Accession number', '')
        u['correspondence'] = rf.get('Corresponding author(s)', '')
        u['article_number'] = rf.get('Article number', '')
        u['abbreviated_source'] = rf.get('Abbreviated source title', '')
        u['eissn'] = rf.get('E-ISSN', '')
        u['conference_name'] = rf.get('Conference name', '')
        u['conference_date'] = rf.get('Conference date', '')
        u['conference_location'] = rf.get('Conference location', '')
        u['conference_code'] = rf.get('Conference code', '')
        u['sponsors'] = rf.get('Sponsor', '')
        u['coden'] = rf.get('CODEN', '')
        u['classification_code'] = rf.get('Classification code', '')
        u['main_heading'] = rf.get('Main heading', '')
        u['num_references'] = rf.get('Number of references', '')
        u['funding_text'] = rf.get('Funding text', '')
        u['open_access'] = rf.get('Open Access type(s)', '')
        u['publication_date'] = rf.get('Issue date', '')

    elif rec.source_db == 'ei' and rec.source_format == 'txt':
        u['issue'] = rf.get('Issue', '')
        u['document_type'] = rf.get('Document type', '')
        u['language'] = rf.get('Language', '')
        u['publisher'] = rf.get('Publisher', '')
        u['affiliations'] = rf.get('Author affiliation', '')
        u['funding'] = rf.get('Funding details', '')
        u['index_keywords'] = rf.get('Controlled terms', '')
        u['authors_raw_short'] = rec.authors
        # Extended fields
        u['accession_number'] = rf.get('Accession number', '')
        u['correspondence'] = rf.get('Corresponding author', '')
        u['abbreviated_source'] = rf.get('Abbreviated source title', '')
        u['conference_name'] = rf.get('Conference name', '')
        u['conference_date'] = rf.get('Conference date', '')
        u['conference_location'] = rf.get('Conference location', '')
        u['conference_code'] = rf.get('Conference code', '')
        u['sponsors'] = rf.get('Sponsor', '')
        u['coden'] = rf.get('CODEN', '')
        u['classification_code'] = rf.get('Classification code', '')
        u['main_heading'] = rf.get('Main heading', '')
        u['num_references'] = rf.get('Number of references', '')
        u['funding_text'] = rf.get('Funding text', '')
        u['publication_date'] = rf.get('Issue date', '')

    # --- Cross-database field inference ---

    # Infer publication_type from document_type for non-WoS records
    if not u['publication_type']:
        dt = u.get('document_type', '').lower()
        if any(kw in dt for kw in ('book chapter', 'book review', 'book')):
            u['publication_type'] = 'B'
        elif any(kw in dt for kw in ('conference', 'proceedings')):
            u['publication_type'] = 'S'
        else:
            u['publication_type'] = 'J'

    # Derive num_references by counting references if not provided
    if not u['num_references'] and u.get('references'):
        refs = u['references']
        # Split by semicolons (common in all databases), filter out fragments
        ref_list = [r.strip() for r in refs.split(';') if len(r.strip()) > 10]
        if not ref_list:
            # Fallback: split by newlines (some TXT formats)
            ref_list = [r.strip() for r in refs.split('\n') if len(r.strip()) > 10]
        if ref_list:
            u['num_references'] = str(len(ref_list))

    return u


def _merge_unified(existing: dict, rec: dict):
    """Fill empty fields in *existing* from *rec*."""
    for key in UNIFIED_FIELDS:
        if not existing.get(key) and rec.get(key):
            existing[key] = rec[key]
    # Also merge raw fields
    raw_e = existing.setdefault('_raw_fields', {})
    for k, v in rec.get('_raw_fields', {}).items():
        if k not in raw_e or not raw_e[k]:
            raw_e[k] = v


def deduplicate_records(unified_records: List[dict]) -> List[dict]:
    seen_doi = {}
    seen_title = {}
    deduped = []

    for rec in unified_records:
        doi = rec.get('doi', '').strip().lower()
        title = rec.get('title', '').strip().lower()

        if doi and doi in seen_doi:
            _merge_unified(seen_doi[doi], rec)
            continue

        from utils import normalize_title as _nt
        norm_t = _nt(title)
        if norm_t and norm_t in seen_title:
            _merge_unified(seen_title[norm_t], rec)
            continue

        deduped.append(rec)
        if doi:
            seen_doi[doi] = rec
        if norm_t:
            seen_title[norm_t] = rec

    return deduped


# ---------------------------------------------------------------------------
# Author Name Conversion for Merged Export
# ---------------------------------------------------------------------------

def _strip_paren(name: str) -> tuple:
    """Remove trailing parenthetical like (58847090400) or (1,2).
    Returns (clean_name, parenthetical_content)."""
    m = re.match(r'^(.*?)\s*\(([^)]*)\)\s*$', name.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return name.strip(), ''


def _is_abbrev(first: str) -> bool:
    """Check if a first-name string is abbreviated (initials only).
    'S' -> True, 'S.' -> True, 'N.I.' -> True, 'NI' -> True,
    'Shuoting' -> False, 'Nikita Igorevich' -> False
    """
    cleaned = first.replace('.', '').replace(' ', '').strip()
    if not cleaned:
        return False
    return cleaned.isupper() and len(cleaned) <= 5


def _make_initials(first_full: str) -> str:
    """Generate initials from full first name.

    Handles both Western multi-word names and Chinese pinyin names:
      'Nikita Igorevich' -> 'NI'  (two words, one initial each)
      'Jean-Pierre'      -> 'JP'  (hyphenated, one initial each)
      'Yanqi'            -> 'YQ'  (single-word Chinese pinyin: Yan+Qi)
      'Shengli'          -> 'SL'  (single-word Chinese pinyin: Sheng+Li)
      'Sheng'            -> 'S'   (single syllable)
      'Shuoting'         -> 'ST'  (single-word Chinese pinyin: Shuo+Ting)
    """
    if not first_full or not first_full.strip():
        return ''
    parts = re.split(r'[\s\-]+', first_full.strip())
    if len(parts) > 1:
        # Multi-word name: take first letter of each part
        return ''.join(p[0].upper() for p in parts if p and p[0].isalpha())
    # Single-word name: try to split as Chinese pinyin syllables
    word = parts[0]
    syllables = _split_pinyin(word.lower())
    if len(syllables) > 1:
        return ''.join(s[0].upper() for s in syllables)
    return word[0].upper() if word else ''


# Valid Mandarin pinyin syllables (without tones), sorted longest-first
# for greedy matching.
_PINYIN_SYLLABLES = sorted([
    'zhuang', 'chuang', 'shuang', 'xiang', 'zhang', 'zheng', 'zhong',
    'zhuai', 'zhuan', 'chang', 'cheng', 'chong', 'chuai', 'chuan',
    'guang', 'huang', 'jiang', 'jiong', 'kuang', 'liang', 'niang',
    'qiang', 'qiong', 'shang', 'sheng', 'shuai', 'shuan', 'xiang',
    'xiong', 'zhuai', 'zhuan', 'zhong', 'zheng', 'zhang',
    'bang', 'beng', 'bing', 'cang', 'ceng', 'chai', 'chan', 'chao',
    'chen', 'chi', 'chou', 'chua', 'chui', 'chun', 'chuo',
    'cong', 'cuan', 'dang', 'deng', 'dian', 'diao', 'ding', 'dong',
    'duan', 'fang', 'feng', 'fiao', 'gang', 'geng', 'gong', 'guai',
    'guan', 'gui', 'guan', 'hang', 'heng', 'hong', 'huai', 'huan',
    'hun', 'huo', 'jian', 'jiao', 'jing', 'juan', 'jun',
    'kang', 'keng', 'kong', 'kuai', 'kuan', 'lang', 'leng', 'lian',
    'liao', 'ling', 'long', 'luan', 'lun', 'mang', 'meng', 'mian',
    'miao', 'ming', 'nang', 'neng', 'nian', 'niao', 'ning', 'nong',
    'nuan', 'pang', 'peng', 'pian', 'piao', 'ping', 'qian', 'qiao',
    'qing', 'quan', 'rang', 'reng', 'rong', 'ruan', 'sang', 'seng',
    'shan', 'shao', 'shei', 'shen', 'shi', 'shou', 'shua', 'shui',
    'shun', 'shuo', 'song', 'suan', 'tang', 'teng', 'tian', 'tiao',
    'ting', 'tong', 'tuan', 'wang', 'weng', 'xian', 'xiao', 'xing',
    'xuan', 'yang', 'ying', 'yong', 'yuan', 'zang', 'zeng', 'zhai',
    'zhan', 'zhao', 'zhei', 'zhen', 'zhi', 'zhou', 'zhua', 'zhui',
    'zhun', 'zhuo', 'zong', 'zuan', 'zui', 'zun', 'zuo',
    'ai', 'an', 'ao', 'ba', 'bi', 'bo', 'bu', 'ca', 'ce', 'ci',
    'cu', 'da', 'de', 'di', 'du', 'en', 'er', 'fa', 'fo', 'fu',
    'ga', 'ge', 'gu', 'ha', 'he', 'hu', 'ji', 'ju', 'ka', 'ke',
    'ku', 'la', 'le', 'li', 'lo', 'lu', 'lv', 'ma', 'me', 'mi',
    'mo', 'mu', 'na', 'ne', 'ni', 'nu', 'nv', 'ou', 'pa', 'pi',
    'po', 'pu', 'qi', 'qu', 're', 'ri', 'ru', 'sa', 'se', 'si',
    'su', 'ta', 'te', 'ti', 'tu', 'wa', 'wo', 'wu', 'xi', 'xu',
    'ya', 'ye', 'yi', 'yu', 'za', 'ze', 'zi', 'zu',
    'a', 'e', 'o',
    # Additional less common but valid
    'bai', 'ban', 'bao', 'bei', 'ben', 'bia', 'bin', 'can', 'cao',
    'cen', 'cha', 'che', 'chu', 'cou', 'cui', 'cun', 'cuo', 'dai',
    'dan', 'dao', 'dei', 'den', 'dia', 'die', 'diu', 'dou', 'dui',
    'dun', 'duo', 'fan', 'fei', 'fen', 'fou', 'gai', 'gan', 'gao',
    'gei', 'gen', 'gou', 'gua', 'gui', 'gun', 'guo', 'hai', 'han',
    'hao', 'hei', 'hen', 'hou', 'hua', 'hui', 'hun', 'huo', 'jia',
    'jie', 'jin', 'jiu', 'jue', 'jun', 'kai', 'kan', 'kao', 'ken',
    'kou', 'kua', 'kui', 'kun', 'kuo', 'lai', 'lan', 'lao', 'lei',
    'lia', 'lie', 'lin', 'liu', 'lou', 'luo', 'mai', 'man', 'mao',
    'mei', 'men', 'min', 'miu', 'mou', 'nai', 'nan', 'nao', 'nei',
    'nie', 'niu', 'nou', 'pai', 'pan', 'pao', 'pei', 'pen', 'pie',
    'pin', 'pou', 'qia', 'qie', 'qin', 'qiu', 'que', 'qun', 'ran',
    'rao', 'ren', 'rou', 'rui', 'run', 'ruo', 'sai', 'san', 'sao',
    'sen', 'sha', 'she', 'shu', 'sou', 'sui', 'sun', 'suo', 'tai',
    'tan', 'tao', 'tie', 'tou', 'tui', 'tun', 'tuo', 'wai', 'wan',
    'wei', 'wen', 'xia', 'xie', 'xin', 'xiu', 'xue', 'xun', 'yan',
    'yao', 'yin', 'you', 'yue', 'yun', 'zai', 'zan', 'zao', 'zei',
    'zen', 'zha', 'zhe', 'zhu', 'zou',
], key=len, reverse=True)


def _split_pinyin(text: str) -> list:
    """Attempt to split a lowercase string into valid pinyin syllables.
    Uses greedy longest-match. Returns a list of syllables,
    or the original text as a single-element list if splitting fails."""
    if not text:
        return []
    remaining = text.lower()
    result = []
    while remaining:
        matched = False
        for syl in _PINYIN_SYLLABLES:
            if remaining.startswith(syl):
                result.append(syl)
                remaining = remaining[len(syl):]
                matched = True
                break
        if not matched:
            # Not a valid pinyin sequence – return original as single element
            return [text]
    return result if len(result) > 1 else [text]


def _initials_with_periods(abbrev: str) -> str:
    """Format initials with periods: 'NI' -> 'N.I.', 'S' -> 'S.'"""
    clean = abbrev.replace('.', '').replace(' ', '').strip()
    if not clean:
        return ''
    return '.'.join(list(clean)) + '.'


def _split_authors(s: str) -> list:
    """Split an author string into individual name strings.
    Handles semicolon, newline, and Scopus comma-separated formats."""
    if not s or not s.strip():
        return []
    s = s.strip()

    # Semicolons are the most common separator
    if ';' in s:
        return [n.strip() for n in s.split(';') if n.strip()]

    # WoS TXT uses newlines
    if '\n' in s:
        return [n.strip() for n in s.split('\n') if n.strip()]

    # Scopus TXT short form: "Last, I., Last, I." pattern
    # Split after "Initial." + ", " + next uppercase letter
    if re.search(r',\s+[A-Z]\.', s):
        parts = re.split(r'(?<=\.),\s+(?=[A-Z])', s)
        if len(parts) > 1:
            return [p.strip() for p in parts if p.strip()]

    # Single author or unknown format
    return [s.strip()] if s.strip() else []


def _parse_name(name_str: str) -> tuple:
    """Parse a single author name string into (last, first, paren_content)."""
    clean, paren = _strip_paren(name_str)
    if not clean:
        return '', '', paren

    if ',' in clean:
        parts = clean.split(',', 1)
        return parts[0].strip(), parts[1].strip(), paren

    # No comma: "Firstname Lastname" format
    parts = clean.split()
    if len(parts) > 1:
        return parts[-1].strip(), ' '.join(parts[:-1]).strip(), paren
    return clean, '', paren


def _parse_all_authors(unified: dict) -> list:
    """Parse authors from a unified record into a structured list.

    Returns list of dicts with keys:
      last, first_full, first_abbrev, scopus_id, affil_num
    """
    source_db = unified.get('source_db', '')
    authors_raw_short = unified.get('authors_raw_short', '')
    author_full_str = unified.get('author_full', '')
    authors_str = unified.get('authors', '')

    # Split each available source into individual name strings
    short_list = _split_authors(authors_raw_short) if authors_raw_short else []
    full_list = _split_authors(author_full_str) if author_full_str else []
    main_list = _split_authors(authors_str) if authors_str else []

    # Determine primary (full names) and secondary (short/abbreviated) lists
    if full_list and short_list:
        primary, secondary = full_list, short_list
    elif full_list:
        secondary_candidate = main_list if (main_list and main_list != full_list) else []
        primary, secondary = full_list, secondary_candidate
    elif short_list:
        secondary_candidate = main_list if (main_list and main_list != short_list) else []
        primary, secondary = short_list, secondary_candidate
    elif main_list:
        primary, secondary = main_list, []
    else:
        return []

    result = []
    n = max(len(primary), len(secondary))

    for i in range(n):
        author = {
            'last': '', 'first_full': '', 'first_abbrev': '',
            'scopus_id': '', 'affil_num': '',
        }

        # Parse from primary list
        if i < len(primary):
            last, first, paren = _parse_name(primary[i])
            author['last'] = last
            if first:
                if _is_abbrev(first):
                    author['first_abbrev'] = first.replace('.', '').strip()
                else:
                    author['first_full'] = first
                    author['first_abbrev'] = _make_initials(first)
            if paren:
                if source_db == 'scopus' and re.match(r'^\d{5,}', paren):
                    author['scopus_id'] = paren
                elif source_db == 'ei':
                    author['affil_num'] = paren

        # Merge from secondary list
        if i < len(secondary):
            last, first, paren = _parse_name(secondary[i])
            if not author['last']:
                author['last'] = last
            if first:
                if _is_abbrev(first):
                    # Prefer explicit abbreviation from source over auto-generated
                    # e.g. WoS AU 'YQ' is more accurate than _make_initials('Yanqi')='Y'
                    author['first_abbrev'] = first.replace('.', '').strip()
                else:
                    if not author['first_full']:
                        author['first_full'] = first
                    if not author['first_abbrev']:
                        author['first_abbrev'] = _make_initials(first)
            if paren:
                if source_db == 'scopus' and re.match(r'^\d{5,}', paren) and not author['scopus_id']:
                    author['scopus_id'] = paren
                elif source_db == 'ei' and not author['affil_num']:
                    author['affil_num'] = paren

        if author['last']:
            result.append(author)

    return result


def _format_for_wos(parsed_authors: list) -> tuple:
    """Format parsed authors for WoS output.
    WoS uses 'Last, Initials' (no periods) for AU and 'Last, FullName' for AF.
    Returns (short_str, full_str), both semicolon-separated.
    """
    short_parts = []
    full_parts = []
    for a in parsed_authors:
        last = a['last']
        initials = a['first_abbrev'] or _make_initials(a['first_full'])
        initials = initials.replace('.', '')  # WoS: no periods on initials
        first_full = a['first_full'] or initials  # Fallback to initials if no full name

        short_parts.append(f'{last}, {initials}' if initials else last)
        full_parts.append(f'{last}, {first_full}' if first_full else last)

    return '; '.join(short_parts), '; '.join(full_parts)


def _format_for_scopus(parsed_authors: list, target_format: str = 'csv') -> tuple:
    """Format parsed authors for Scopus output.
    Scopus uses 'Last, I.N.' (periods on initials) for short form and
    'Last, FullName' for full form.  CSV full form includes Scopus IDs.
    Returns (short_str, full_str), both semicolon-separated.
    """
    short_parts = []
    full_parts = []
    for a in parsed_authors:
        last = a['last']
        initials = a['first_abbrev'] or _make_initials(a['first_full'])
        init_dotted = _initials_with_periods(initials) if initials else ''

        short_parts.append(f'{last}, {init_dotted}' if init_dotted else last)

        first_full = a['first_full'] or init_dotted or ''
        name = f'{last}, {first_full}' if first_full else last
        # Scopus IDs only appear in CSV Author full names, not in TXT
        if target_format == 'csv' and a.get('scopus_id'):
            name += f' ({a["scopus_id"]})'
        full_parts.append(name)

    return '; '.join(short_parts), '; '.join(full_parts)


def _format_for_ei(parsed_authors: list) -> str:
    """Format parsed authors for EI output.
    EI uses 'Last, FullName (AffiliationNum)' semicolon-separated.
    Returns authors_str.
    """
    parts = []
    for a in parsed_authors:
        last = a['last']
        first = a['first_full'] or a['first_abbrev'] or ''
        name = f'{last}, {first}' if first else last
        if a.get('affil_num'):
            name += f' ({a["affil_num"]})'
        parts.append(name)
    return '; '.join(parts)


# ---------------------------------------------------------------------------
# Cross-database affiliation / correspondence format conversion
# ---------------------------------------------------------------------------

def _convert_ei_affil_to_wos_c1(unified: dict) -> str:
    """Convert EI numbered affiliations to WoS C1 format.

    EI input:
      Authors: Gu, Sheng (1, 3); Wu, Yanqi (1, 2); Wang, Xidong (2)
      Affiliations: (1) Southeast Univ, Nanjing, China; (2) Zhengzhou Univ, ...

    WoS C1 output:
      [Gu, Sheng; Wu, Yanqi] Southeast Univ, Nanjing, China.
      [Wu, Yanqi; Wang, Xidong] Zhengzhou Univ, ...
    """
    authors_str = unified.get('authors_raw_short', '') or unified.get('authors', '')
    affil_str = unified.get('affiliations', '')
    if not authors_str or not affil_str:
        return affil_str

    # Parse numbered affiliations: (1) Institution...; (2) Institution...
    affil_map = {}
    for m in re.finditer(r'\((\d+)\)\s*((?:(?!\(\d+\)).)*)', affil_str):
        num = m.group(1)
        text = m.group(2).strip().rstrip(';').strip()
        if text:
            affil_map[num] = text

    if not affil_map:
        return affil_str

    # Parse authors with numbers: Name (1, 3); Name (1, 2)
    num_to_authors = {}
    for entry in authors_str.split(';'):
        entry = entry.strip()
        m = re.match(r'^(.*?)\s*\(([0-9,\s]+)\)\s*$', entry)
        if m:
            name = m.group(1).strip()
            # Strip affil numbers, keep just the name
            name_clean = re.sub(r'\s*\([0-9,\s]+\)\s*$', '', name).strip()
            nums = [n.strip() for n in m.group(2).split(',')]
            for num in nums:
                num_to_authors.setdefault(num, []).append(name_clean)

    # Build WoS C1 format: [Author1; Author2] Institution.
    parts = []
    for num in sorted(affil_map.keys(), key=lambda x: int(x)):
        inst = affil_map[num]
        authors = num_to_authors.get(num, [])
        if authors:
            author_list = '; '.join(authors)
            parts.append(f'[{author_list}] {inst}.')
        else:
            parts.append(f'{inst}.')

    return '\n   '.join(parts)


def _convert_scopus_affil_to_wos_c1(unified: dict) -> str:
    """Convert Scopus affiliations to WoS C1 format.

    Uses 'authors_with_affiliations' field which has:
      Xiao, Shuoting, Institute of..., City, Country;
      Fomin, Nikita Igorevich, Institute of..., City, Country

    WoS C1 output:
      [Xiao, Shuoting; Fomin, Nikita Igorevich] Institute of..., City, Country.
    """
    awa = unified.get('authors_with_affiliations', '')
    if not awa:
        return unified.get('affiliations', '')

    # Get known author names to split name from affiliation
    full_names = _split_authors(
        unified.get('author_full', '') or unified.get('authors', '')
    )
    # Clean names (remove Scopus IDs, affil numbers)
    clean_names = []
    for n in full_names:
        n = re.sub(r'\s*\(\d[\d,\s]*\)\s*$', '', n).strip()  # EI nums
        n = re.sub(r'\s*\(\d{5,}\)\s*$', '', n).strip()       # Scopus IDs
        clean_names.append(n)

    # Parse each "Author, Affiliation" entry
    author_affil_pairs = []
    for entry in awa.split(';'):
        entry = entry.strip()
        if not entry:
            continue

        matched = False
        for known_name in clean_names:
            if entry.startswith(known_name):
                affil = entry[len(known_name):].lstrip(',').strip()
                author_affil_pairs.append((known_name, affil))
                matched = True
                break

        if not matched:
            # Fallback: assume "Last, First, Affiliation..."
            parts = entry.split(',', 2)
            if len(parts) >= 3:
                name = f'{parts[0].strip()}, {parts[1].strip()}'
                affil = parts[2].strip()
                author_affil_pairs.append((name, affil))

    if not author_affil_pairs:
        return unified.get('affiliations', '')

    # Group authors by affiliation
    from collections import OrderedDict
    affil_to_authors = OrderedDict()
    for name, affil in author_affil_pairs:
        affil_to_authors.setdefault(affil, []).append(name)

    # Build WoS C1 format
    parts = []
    for affil, authors in affil_to_authors.items():
        author_list = '; '.join(authors)
        parts.append(f'[{author_list}] {affil}.')

    return '\n   '.join(parts)


def _convert_affil_to_wos_c1(unified: dict) -> str:
    """Convert affiliations from any source database to WoS C1 format."""
    source_db = unified.get('source_db', '')
    if source_db == 'wos':
        return unified.get('affiliations', '')
    elif source_db == 'ei':
        return _convert_ei_affil_to_wos_c1(unified)
    elif source_db == 'scopus':
        return _convert_scopus_affil_to_wos_c1(unified)
    return unified.get('affiliations', '')


def _convert_corresp_to_wos_rp(unified: dict, parsed_authors: list) -> str:
    """Convert correspondence/corresponding author to WoS RP format.

    WoS RP format: Li, SL (通讯作者)，Zhengzhou Univ, Sch Civil Engn, Zhengzhou 450001, Peoples R China.

    Scopus format: S. Xiao; Institute of..., City, Country; email: xxx
    EI format: Li, Shengli(lsl@zzu.edu.cn)
    """
    source_db = unified.get('source_db', '')
    corresp = unified.get('correspondence', '') or unified.get('reprint_address', '')
    if not corresp or source_db == 'wos':
        return corresp

    email = ''
    name_abbrev = ''
    institution = ''

    if source_db == 'scopus':
        # Scopus: "S. Xiao; Institution, City, Country; email: xxx"
        parts = corresp.split(';')
        name_part = parts[0].strip()
        inst_parts = []
        for p in parts[1:]:
            p = p.strip()
            if p.lower().startswith('email:'):
                email = p[6:].strip()
            else:
                inst_parts.append(p)
        institution = ', '.join(inst_parts).strip().rstrip('.')

        # Convert Scopus name "S. Xiao" to WoS format "Xiao, S"
        name_tokens = name_part.replace('.', '. ').split()
        if len(name_tokens) >= 2:
            # Check if first token(s) are initials, last token is surname
            initials = []
            surname = ''
            for t in name_tokens:
                t_clean = t.replace('.', '').strip()
                if t_clean and len(t_clean) <= 2 and t_clean[0].isupper():
                    initials.append(t_clean)
                else:
                    surname = t.strip()
            if surname and initials:
                name_abbrev = f'{surname}, {"".join(initials)}'
            else:
                name_abbrev = name_part
        else:
            name_abbrev = name_part

    elif source_db == 'ei':
        # EI: "Li, Shengli(lsl@zzu.edu.cn)"
        m = re.match(r'^(.*?)\s*\(([^)]+@[^)]+)\)\s*$', corresp)
        if m:
            full_name = m.group(1).strip()
            email = m.group(2).strip()
        else:
            full_name = corresp.strip()

        # Convert full name to abbreviated: "Li, Shengli" -> "Li, SL"
        if ',' in full_name:
            last, first = full_name.split(',', 1)
            last = last.strip()
            first = first.strip()
            initials = _make_initials(first) if not _is_abbrev(first) else first.replace('.', '')
            name_abbrev = f'{last}, {initials}'
        else:
            name_abbrev = full_name

        # Try to find institution from affiliations
        affil_str = unified.get('affiliations', '')
        if affil_str and full_name:
            # EI: find which affiliation the corresponding author belongs to
            for entry in (unified.get('authors_raw_short', '') or unified.get('authors', '')).split(';'):
                entry = entry.strip()
                m2 = re.match(r'^(.*?)\s*\(([0-9,\s]+)\)\s*$', entry)
                if m2:
                    entry_name = m2.group(1).strip()
                    # Match by surname
                    if full_name.split(',')[0].strip().lower() == entry_name.split(',')[0].strip().lower():
                        first_num = m2.group(2).split(',')[0].strip()
                        # Find this affiliation number
                        m3 = re.search(r'\(' + first_num + r'\)\s*((?:(?!\(\d+\)).)*)', affil_str)
                        if m3:
                            institution = m3.group(1).strip().rstrip(';').strip()
                        break

    if not name_abbrev:
        return corresp

    rp = f'{name_abbrev} (通讯作者)，{institution}.' if institution else f'{name_abbrev} (通讯作者)'

    # Also set email if we extracted one
    if email:
        unified.setdefault('_extracted_email', email)

    return rp


def _convert_scopus_ref_to_wos_cr(ref_entry: str) -> str:
    """Convert a single Scopus reference entry to WoS CR format.

    Scopus: 'Blismas, Nick G., Benefit evaluation for off-site production in construction,
             Construction Management and Economics, 24, 2, pp. 121-130, (2006)'
    WoS CR: 'Blismas NG, 2006, CONSTR MANAGE ECON, V24, P121, DOI ...'

    Since Scopus references don't always contain DOI and we can't reliably
    abbreviate journal names, we produce a best-effort conversion:
      AuthorLastname Initials, Year, JOURNAL TITLE, VVolume, PFirstPage
    """
    ref_entry = ref_entry.strip()
    if not ref_entry:
        return ''

    # Extract year: look for (YYYY) at or near the end
    year = ''
    year_match = re.search(r'\((\d{4})\)\s*$', ref_entry)
    if year_match:
        year = year_match.group(1)
        # Remove the year portion from the entry for further parsing
        ref_body = ref_entry[:year_match.start()].rstrip(', ')
    else:
        ref_body = ref_entry

    # Split by commas to parse fields
    # Scopus format: Author(s), Article Title, Journal, Volume, Issue, pp. Pages
    parts = [p.strip() for p in ref_body.split(',')]

    # Extract author name (first part before article title)
    # Author format: "LastName, FirstName" or "LastName, Initials"
    author_wos = ''
    title_start_idx = 0

    if len(parts) >= 2:
        # Check if parts[0] looks like a surname and parts[1] like a first name
        potential_last = parts[0].strip()
        potential_first = parts[1].strip()

        # Heuristic: author name has short first part (surname) and
        # second part starts with uppercase (given name or initials)
        if (potential_first and potential_first[0].isupper() and
                len(potential_last.split()) <= 3 and
                not potential_last.startswith('pp.') and
                not re.match(r'^\d', potential_last)):
            # This looks like an author: "LastName, FirstName MiddleName"
            first_full = potential_first
            # Generate WoS-style initials
            first_parts = re.split(r'[\s\-]+', first_full)
            initials = ''.join(p[0].upper() for p in first_parts if p and p[0].isalpha())
            # Try pinyin splitting for single-word names
            if len(first_parts) == 1:
                syllables = _split_pinyin(first_parts[0].lower())
                if len(syllables) > 1:
                    initials = ''.join(s[0].upper() for s in syllables)
            author_wos = f'{potential_last} {initials}'
            title_start_idx = 2
        else:
            # No clear author — might be an anonymous reference
            title_start_idx = 0
    else:
        title_start_idx = 0

    # Find volume (a standalone number or number after journal name)
    volume = ''
    first_page = ''
    journal = ''

    # Look for "pp. NNN-NNN" or "pp. NNN" for pages
    remaining_parts = parts[title_start_idx:]
    page_idx = -1
    for idx, p in enumerate(remaining_parts):
        pp_match = re.match(r'^pp\.\s*(\d+)', p)
        if pp_match:
            first_page = pp_match.group(1)
            page_idx = idx
            break

    # Look for volume and issue (standalone numbers before pp.)
    # Pattern: ..., Journal Name, Volume, Issue, pp. Pages
    # We scan backwards from pages (or end) to find volume
    search_end = page_idx if page_idx >= 0 else len(remaining_parts)

    # Find the last numeric-only part before pages — that's likely Issue
    # The one before that is Volume
    num_indices = []
    for idx in range(search_end):
        if remaining_parts[idx].strip().isdigit():
            num_indices.append(idx)

    if len(num_indices) >= 2:
        # Volume and Issue found
        vol_idx = num_indices[-2]
        volume = remaining_parts[vol_idx].strip()
        # Journal is everything between title and volume
        journal_parts = remaining_parts[:vol_idx]
    elif len(num_indices) == 1:
        vol_idx = num_indices[0]
        volume = remaining_parts[vol_idx].strip()
        journal_parts = remaining_parts[:vol_idx]
    else:
        # No volume found — try to identify the journal from remaining parts
        # Skip the article title (usually the longest part) and take the next
        journal_parts = remaining_parts
        vol_idx = len(remaining_parts)

    # The article title is typically the first part(s) after author,
    # and the journal name is the part just before the volume number
    # In Scopus: Author, Article Title, Journal Name, Vol, Issue, pp. Pages
    # We want just the journal name for WoS CR
    if journal_parts:
        # The journal is typically the last entry before volume
        # Article title is everything else
        if len(journal_parts) >= 2:
            # Last non-empty part before volume is likely the journal
            journal = journal_parts[-1].strip()
        elif journal_parts:
            journal = journal_parts[0].strip()

    # Build WoS CR format
    cr_parts = []
    if author_wos:
        cr_parts.append(author_wos)
    if year:
        cr_parts.append(year)
    if journal:
        cr_parts.append(journal.upper())
    if volume:
        cr_parts.append(f'V{volume}')
    if first_page:
        cr_parts.append(f'P{first_page}')

    return ', '.join(cr_parts) if cr_parts else ref_entry


def _convert_refs_to_wos_cr(unified: dict) -> str:
    """Convert references from Scopus format to WoS CR format.

    Scopus refs are semicolon-separated with full author names, article titles,
    and full journal names. WoS CR uses abbreviated form:
      AuthorLast Initials, Year, JOURNAL ABBREV, VVolume, PPage, DOI xxx

    EI exports typically don't contain reference lists, so only Scopus is handled.
    """
    source_db = unified.get('source_db', '')
    refs = unified.get('references', '')
    if not refs or source_db == 'wos':
        return refs  # WoS refs already in CR format

    if source_db != 'scopus':
        return refs  # EI has no refs to convert; leave as-is

    # Split Scopus references by semicolons
    # Be careful: some entries have semicolons inside parenthetical notes
    ref_entries = refs.split(';')
    converted = []
    for entry in ref_entries:
        entry = entry.strip()
        if not entry or len(entry) < 10:
            continue
        cr = _convert_scopus_ref_to_wos_cr(entry)
        if cr:
            converted.append(cr)

    return '; '.join(converted) if converted else refs


def _generate_wos_ut(unified: dict) -> str:
    """Generate a UT identifier for records from non-WoS databases.

    For Scopus records: 'SCOPUS:2-s2.0-XXXXX' using the EID
    For EI records: 'COMPENDEX:XXXXX' using the accession number
    WoS records already have their own UT.
    """
    source_db = unified.get('source_db', '')
    if source_db == 'wos':
        return unified.get('wos_id', '')

    accession = unified.get('accession_number', '')

    if source_db == 'scopus':
        # Try to get EID from accession_number or from link URL
        eid = accession
        if not eid:
            link = unified.get('link', '')
            m = re.search(r'eid=(2-s2\.0-\d+)', link)
            if m:
                eid = m.group(1)
        if eid:
            # Ensure it has the standard prefix
            if eid.startswith('2-s2.0-'):
                return f'SCOPUS:{eid}'
            else:
                return f'SCOPUS:{eid}'
        return ''

    elif source_db == 'ei':
        if accession:
            return f'COMPENDEX:{accession}'
        return ''

    return ''


def _convert_authors_for_export(
    unified_records: List[dict],
    target_db: str,
    target_format: str,
) -> List[dict]:
    """Convert author fields and other database-specific fields in all unified
    records to match the target database format style.

    Args:
        unified_records: Unified record dicts from record_to_unified().
        target_db: Target database ('wos', 'scopus', 'ei').
        target_format: Target file format ('txt', 'csv', 'xls').

    Returns:
        New list of dicts with converted fields (originals unchanged).
    """
    if not target_db:
        return unified_records

    result = []
    for rec in unified_records:
        rec_copy = dict(rec)
        parsed = _parse_all_authors(rec_copy)
        if parsed:
            if target_db == 'wos':
                short_str, full_str = _format_for_wos(parsed)
                rec_copy['authors'] = short_str
                rec_copy['author_full'] = full_str
            elif target_db == 'scopus':
                short_str, full_str = _format_for_scopus(parsed, target_format)
                rec_copy['authors'] = short_str
                rec_copy['author_full'] = full_str
            elif target_db == 'ei':
                rec_copy['authors'] = _format_for_ei(parsed)
                rec_copy['author_full'] = ''

        # Convert affiliations and correspondence for WoS target
        if target_db == 'wos' and rec_copy.get('source_db') != 'wos':
            # Convert correspondence FIRST (needs original affiliations format
            # to look up corresponding author's institution)
            new_rp = _convert_corresp_to_wos_rp(rec_copy, parsed or [])
            if new_rp:
                rec_copy['reprint_address'] = new_rp
                rec_copy['correspondence'] = new_rp

            # Extract email from correspondence if not already set
            extracted_email = rec_copy.pop('_extracted_email', '')
            if extracted_email and not rec_copy.get('email'):
                rec_copy['email'] = extracted_email

            # Convert affiliations to WoS C1 format (after RP uses originals)
            new_c1 = _convert_affil_to_wos_c1(rec_copy)
            if new_c1:
                rec_copy['affiliations'] = new_c1

            # Convert references to WoS CR format
            new_cr = _convert_refs_to_wos_cr(rec_copy)
            if new_cr:
                rec_copy['references'] = new_cr

            # Generate UT identifier for non-WoS records
            if not rec_copy.get('wos_id'):
                ut = _generate_wos_ut(rec_copy)
                if ut:
                    rec_copy['wos_id'] = ut

        result.append(rec_copy)
    return result


# ---------------------------------------------------------------------------
# Template-based export
# ---------------------------------------------------------------------------

def _analyze_template_csv(file_path: str) -> Tuple[List[str], object, str]:
    encoding = detect_encoding(file_path)
    dialect = detect_csv_dialect(file_path, encoding)
    with open(file_path, 'r', encoding=encoding, errors='replace', newline='') as f:
        reader = csv.reader(f, dialect)
        headers = next(reader, [])
    return [h.strip() for h in headers], dialect, encoding


def _map_unified_to_template_field(template_header: str, unified: dict) -> str:
    mapping = {
        # Title
        'Title': 'title', '文献标题': 'title', 'Article Title': 'title',
        # Authors
        'Authors': 'authors', '作者': 'authors', 'Author': 'authors',
        'Author full names': 'author_full', 'Author Full Names': 'author_full',
        # Year
        'Year': 'year', '年份': 'year',
        'Publication year': 'year', 'Publication Year': 'year',
        # Journal / Source
        'Source title': 'journal', '来源出版物名称': 'journal',
        'Source Title': 'journal', 'Source': 'journal', 'SO': 'journal',
        'Abbreviated source title': 'abbreviated_source',
        'Abbreviated Source Title': 'abbreviated_source',
        '来源出版物名称缩写': 'abbreviated_source',
        'Journal Abbreviation': 'journal_abbreviation',
        'Journal ISO Abbreviation': 'journal_iso',
        # Volume / Issue / Pages
        'Volume': 'volume', '卷': 'volume',
        'Issue': 'issue', '期': 'issue',
        'Pages': 'pages',
        'Page start': 'page_start', '起始页码': 'page_start',
        'Page end': 'page_end', '结束页码': 'page_end',
        'Start Page': 'page_start', 'End Page': 'page_end',
        'Page Count': 'page_count',
        'Art. No.': 'article_number', '文献编号': 'article_number',
        'Article Number': 'article_number', 'Article number': 'article_number',
        # Identifiers
        'DOI': 'doi',
        'ISSN': 'issn', 'ISBN': 'isbn', 'ISBN13': 'isbn', 'ISBN-13': 'isbn',
        'eISSN': 'eissn', 'E-ISSN': 'eissn',
        'CODEN': 'coden',
        'PubMed ID': 'pubmed_id',
        'EID': 'accession_number',
        'UT (Unique WOS ID)': 'wos_id',
        'Accession number': 'accession_number',
        # Content
        'Abstract': 'abstract', '摘要': 'abstract',
        'Author Keywords': 'author_keywords', '作者关键字': 'author_keywords',
        'Uncontrolled terms': 'author_keywords',
        'Index Keywords': 'index_keywords', '索引关键字': 'index_keywords',
        'Keywords Plus': 'index_keywords',
        'Controlled/Subject terms': 'index_keywords',
        'Controlled terms': 'index_keywords',
        'Main heading': 'main_heading',
        'Classification code': 'classification_code',
        # Document metadata
        'Document Type': 'document_type', '文献类型': 'document_type',
        'Document type': 'document_type',
        'Publication Type': 'publication_type',
        'Language': 'language',
        'Language of Original Document': 'language', '原始文献语言': 'language',
        'Publication Stage': 'publication_stage', '出版阶段': 'publication_stage',
        'Open Access': 'open_access', '开放获取': 'open_access',
        'Open Access Designations': 'open_access',
        'Open Access type(s)': 'open_access',
        'Publication Date': 'publication_date',
        'Issue date': 'publication_date',
        # Citations
        'Cited by': 'cited_by', '引用次数': 'cited_by',
        'Times Cited, All Databases': 'cited_by',
        'Times Cited, WoS Core': 'cited_by',
        'Number of Cited References': 'num_references',
        'Number of references': 'num_references',
        'References': 'references', '参考文献': 'references',
        'Cited References': 'references',
        # Publisher
        'Publisher': 'publisher', '出版商': 'publisher',
        'Publisher/Repository': 'publisher',
        'Publisher Address': 'publisher_address',
        'Publisher City': 'publisher_city',
        # Affiliations / Correspondence
        'Affiliations': 'affiliations', '归属机构': 'affiliations',
        'Author affiliation': 'affiliations', 'Addresses': 'affiliations',
        'Authors with affiliations': 'authors_with_affiliations',
        '含归属机构的作者': 'authors_with_affiliations',
        'Correspondence Address': 'correspondence', '通讯地址': 'correspondence',
        'Corresponding author(s)': 'correspondence',
        'Corresponding author': 'correspondence',
        'Reprint Addresses': 'reprint_address',
        'Email Addresses': 'email',
        # Funding
        'Funding Details': 'funding', '出资详情': 'funding',
        'Funding details': 'funding', 'Funding Orgs': 'funding',
        'Funding Text': 'funding_text', '出资文本': 'funding_text',
        'Funding Texts': 'funding_text', 'Funding text': 'funding_text',
        # Author IDs
        'Author(s) ID': 'author_ids', '作者标识号': 'author_ids',
        'ResearcherID Number': 'researcher_ids',
        # Conference
        'Conference name': 'conference_name', '会议名称': 'conference_name',
        'Conference date': 'conference_date', '会议日期': 'conference_date',
        'Conference location': 'conference_location', '会议地点': 'conference_location',
        'Conference code': 'conference_code', '会议代码': 'conference_code',
        'Sponsors': 'sponsors', '主办方': 'sponsors', 'Sponsor': 'sponsors',
        'Editors': 'editors', '编辑': 'editors', 'Book Editors': 'editors',
        # Links
        'Link': 'link', '链接': 'link',
        # Research classification
        'Research Areas': 'research_areas',
        'Web of Science Categories': 'research_areas',
        # WoS-specific
        'Usage Count (Last 180 Days)': 'usage_count_180',
        'Usage Count (Since 2013)': 'usage_count_since2013',
    }
    field_key = mapping.get(template_header, '')
    if field_key:
        val = unified.get(field_key, '')
        if val:
            return val

    # Fallback: try raw fields pass-through (exact header name match)
    raw = unified.get('_raw_fields', {})
    if template_header in raw:
        return raw[template_header]

    return ''


def export_merged_csv(
    unified_records: List[dict],
    template_path: str,
    output_path: str,
    template_db: str = '',
):
    headers, dialect, encoding = _analyze_template_csv(template_path)
    safe = _safe_dialect(dialect)

    converted = _convert_authors_for_export(unified_records, template_db, 'csv')

    with open(output_path, 'w', encoding=encoding, errors='replace', newline='') as f:
        writer = csv.writer(f, dialect=safe)
        writer.writerow(headers)
        for rec in converted:
            row = [_map_unified_to_template_field(h, rec) for h in headers]
            writer.writerow(row)


def export_merged_xls(
    unified_records: List[dict],
    template_path: str,
    output_path: str,
    template_db: str = '',
):
    wb_src = xlrd.open_workbook(template_path)
    ws_src = wb_src.sheet_by_index(0)
    headers = [str(ws_src.cell_value(0, c)).strip() for c in range(ws_src.ncols)]

    converted = _convert_authors_for_export(unified_records, template_db, 'xls')

    try:
        wb_out = xlwt.Workbook(encoding='utf-8')
        ws_out = wb_out.add_sheet('Sheet1')
        for c, h in enumerate(headers):
            ws_out.write(0, c, h)
        for r, rec in enumerate(converted, start=1):
            for c, h in enumerate(headers):
                val = _map_unified_to_template_field(h, rec)
                ws_out.write(r, c, val)
        wb_out.save(output_path)
    except Exception:
        if not output_path.endswith('.xlsx'):
            output_path = output_path.rsplit('.', 1)[0] + '.xlsx'
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.append(headers)
        for rec in converted:
            row = [_map_unified_to_template_field(h, rec) for h in headers]
            ws_out.append(row)
        wb_out.save(output_path)

    return output_path


def _wos_multiline(tag: str, text: str, lines: list, wrap: int = 70):
    """Append a WoS field with word-wrapping continuation lines.

    If the text already contains newlines (e.g. from C1 affiliation conversion),
    each line is treated as a separate entry and word-wrapped individually.
    """
    # Split into logical entries (e.g., separate affiliations)
    entries = text.split('\n')
    first_overall = True
    for entry in entries:
        entry = entry.strip()
        if not entry:
            continue
        # Word-wrap this entry
        words = entry.split()
        wrapped = []
        current = ''
        for w in words:
            if len(current) + len(w) + 1 > wrap and current:
                wrapped.append(current)
                current = w
            else:
                current = (current + ' ' + w).strip()
        if current:
            wrapped.append(current)
        for ln in wrapped:
            if first_overall:
                lines.append(f'{tag} {ln}')
                first_overall = False
            else:
                lines.append(f'   {ln}')


def _wos_name_list(tag: str, name_str: str, lines: list):
    """Append a WoS multi-line name field (AU/AF) with one name per line."""
    names = re.split(r';\s*', name_str)
    for i, name in enumerate(names):
        name = re.sub(r'\s*\(\d+\)\s*$', '', name).strip()
        if not name:
            continue
        lines.append(f'{tag} {name}' if i == 0 else f'   {name}')


def _wos_ref_list(tag: str, refs_str: str, lines: list):
    """Append WoS CR (cited references) field, one reference per line."""
    # Split on semicolons — each reference is a separate entry
    refs = [r.strip() for r in refs_str.split(';') if r.strip()]
    first = True
    for ref in refs:
        if not ref or len(ref) < 5:
            continue
        lines.append(f'{tag} {ref}' if first else f'   {ref}')
        first = False
    if first:
        # No valid refs found
        lines.append(f'{tag} ')


# WoS TXT field definitions in standard output order.
# (tag, unified_key_or_None, format_type, default_or_None)
# format_type: 'single' | 'upper' | 'multi' | 'namelist' | 'reflist'
# For unified_key=None, value comes from raw_fields only.
_WOS_TAG_DEFS = [
    ('PT', 'publication_type', 'single', 'J'),
    ('AU', 'authors', 'namelist', None),
    ('AF', 'author_full', 'namelist', None),
    ('TI', 'title', 'multi', None),
    ('SO', 'journal', 'upper', None),
    ('LA', 'language', 'single', None),
    ('DT', 'document_type', 'single', None),
    ('DE', 'author_keywords', 'multi', None),
    ('ID', 'index_keywords', 'multi', None),
    ('AB', 'abstract', 'multi', None),
    ('C1', 'affiliations', 'multi', None),
    ('C3', 'organization_enhanced', 'multi', None),
    ('RP', 'reprint_address', 'multi', None),   # fallback: correspondence
    ('EM', 'email', 'single', None),
    ('RI', 'researcher_ids', 'multi', None),
    ('OI', None, 'multi', None),
    ('FU', 'funding', 'multi', None),
    ('FX', 'funding_text', 'multi', None),
    ('CR', 'references', 'reflist', None),
    ('NR', 'num_references', 'single', None),
    ('TC', 'cited_by', 'single', None),
    ('Z9', None, 'single', None),
    ('U1', 'usage_count_180', 'single', None),
    ('U2', 'usage_count_since2013', 'single', None),
    ('PU', 'publisher', 'single', None),
    ('PI', 'publisher_city', 'single', None),
    ('PA', 'publisher_address', 'multi', None),
    ('SN', 'issn', 'single', None),
    ('EI', 'eissn', 'single', None),
    ('J9', 'journal_abbreviation', 'single', None),
    ('JI', 'journal_iso', 'single', None),
    ('PD', 'publication_date', 'single', None),
    ('PY', 'year', 'single', None),
    ('VL', 'volume', 'single', None),
    ('IS', 'issue', 'single', None),
    ('BP', 'page_start', 'single', None),
    ('EP', 'page_end', 'single', None),
    ('AR', 'article_number', 'single', None),
    ('DI', 'doi', 'single', None),
    ('PG', 'page_count', 'single', None),
    ('WC', None, 'multi', None),
    ('WE', None, 'single', None),
    ('SC', 'research_areas', 'multi', None),
    ('OA', 'open_access', 'single', None),
    ('EA', None, 'single', None),
    ('GA', None, 'single', None),
    ('UT', 'wos_id', 'single', None),
    ('DA', None, 'single', None),
]


def _build_wos_txt_record(unified: dict) -> str:
    """Build a complete WoS TXT record with ALL standard tags in order.
    Tags are always output even if empty, to match official WoS format."""
    lines = []
    raw = unified.get('_raw_fields', {})

    for tag, ukey, fmt, default in _WOS_TAG_DEFS:
        # Resolve value: unified field -> raw field -> default
        val = ''
        if ukey:
            val = str(unified.get(ukey, '') or '')
        if not val:
            val = str(raw.get(tag, '') or '')
        # Special fallback: RP uses correspondence if reprint_address empty
        if not val and tag == 'RP':
            val = str(unified.get('correspondence', '') or '')
        if not val and default:
            val = default

        # Format and append
        if fmt == 'upper':
            lines.append(f'{tag} {val.upper() if val else ""}')
        elif fmt == 'namelist':
            if val:
                _wos_name_list(tag, val, lines)
            else:
                lines.append(f'{tag} ')
        elif fmt == 'reflist':
            if val:
                _wos_ref_list(tag, val, lines)
            else:
                lines.append(f'{tag} ')
        elif fmt == 'multi':
            if val:
                _wos_multiline(tag, val, lines)
            else:
                lines.append(f'{tag} ')
        else:  # single
            lines.append(f'{tag} {val}')

    lines.append('ER')
    return '\n'.join(lines)


def _build_scopus_txt_record(unified: dict, lang: str = 'english') -> str:
    lines = []

    # First line: short author names, comma-separated
    author_str = unified.get('authors', '')
    if author_str:
        short_names = re.split(r';\s*', author_str)
        short_line = ', '.join(n.strip() for n in short_names if n.strip())
        lines.append(short_line)

    # Author full names (with Scopus IDs if available)
    author_full = unified.get('author_full', '')
    if author_full:
        if lang == 'chinese':
            lines.append(f'作者全名: {author_full}')
        else:
            lines.append(f'AUTHOR FULL NAMES: {author_full}')

    # Title
    title = unified.get('title', '')
    if title:
        lines.append(title)

    # Source line: (year) journal, volume, pages/article_number
    year = unified.get('year', '')
    journal = unified.get('journal', '')
    volume = unified.get('volume', '')
    pages = unified.get('pages', '')
    art_no = unified.get('article_number', '')
    source_parts = []
    if year:
        source_parts.append(f'({year})')
    if journal:
        source_parts.append(journal)
    if volume:
        source_parts.append(volume)
    detail = pages or art_no
    if detail:
        source_parts.append(f', pp. {detail}' if pages else f', {detail}')
    if source_parts:
        lines.append(' '.join(source_parts))

    # Cited by count
    cited_by = unified.get('cited_by', '')
    if cited_by:
        if lang == 'chinese':
            lines.append(f'引用次数: {cited_by}')
        else:
            lines.append(f'CITED BY: {cited_by}')

    # DOI and link
    doi = unified.get('doi', '')
    if doi:
        lines.append(f'DOI: {doi}')
    link = unified.get('link', '')
    if link:
        lines.append(link)

    lines.append('')

    # Build label map for all fields (English / Chinese)
    if lang == 'chinese':
        label_map = [
            ('affiliations', '归属机构'),
            ('abstract', '摘要'),
            ('author_keywords', '作者关键字'),
            ('index_keywords', '索引关键字'),
            ('funding', '出资详情'),
            ('funding_text', '出资文本'),
            ('references', '参考文献'),
            ('correspondence', '通讯地址'),
            ('publisher', '出版商'),
            ('issn', 'ISSN'),
            ('isbn', 'ISBN'),
            ('coden', 'CODEN'),
            ('language', '原始文献语言'),
            ('abbreviated_source', '来源出版物名称缩写'),
            ('document_type', '文献类型'),
            ('publication_stage', '出版阶段'),
            ('open_access', '开放获取'),
        ]
    else:
        label_map = [
            ('affiliations', 'AFFILIATIONS'),
            ('abstract', 'ABSTRACT'),
            ('author_keywords', 'AUTHOR KEYWORDS'),
            ('index_keywords', 'INDEX KEYWORDS'),
            ('funding', 'FUNDING DETAILS'),
            ('funding_text', 'FUNDING TEXT'),
            ('references', 'REFERENCES'),
            ('correspondence', 'CORRESPONDENCE ADDRESS'),
            ('publisher', 'PUBLISHER'),
            ('issn', 'ISSN'),
            ('isbn', 'ISBN'),
            ('coden', 'CODEN'),
            ('language', 'LANGUAGE OF ORIGINAL DOCUMENT'),
            ('abbreviated_source', 'ABBREVIATED SOURCE TITLE'),
            ('document_type', 'DOCUMENT TYPE'),
            ('publication_stage', 'PUBLICATION STAGE'),
            ('open_access', 'OPEN ACCESS'),
        ]

    for field_key, label in label_map:
        val = unified.get(field_key, '')
        if val:
            lines.append(f'{label}: {val}')

    return '\n'.join(lines)


def _build_ei_txt_record(unified: dict, record_num: int) -> str:
    lines = [f'<RECORD {record_num}>', '']

    field_map = [
        ('accession_number', 'Accession number'),
        ('title', 'Title'),
        ('authors', 'Authors'),
        ('affiliations', 'Author affiliation'),
        ('correspondence', 'Corresponding author'),
        ('journal', 'Source title'),
        ('abbreviated_source', 'Abbreviated source title'),
        ('volume', 'Volume'),
        ('issue', 'Issue'),
        ('publication_date', 'Issue date'),
        ('year', 'Publication year'),
        ('pages', 'Pages'),
        ('article_number', 'Article number'),
        ('language', 'Language'),
        ('issn', 'ISSN'),
        ('eissn', 'E-ISSN'),
        ('isbn', 'ISBN-13'),
        ('coden', 'CODEN'),
        ('document_type', 'Document type'),
        ('publisher', 'Publisher'),
        ('abstract', 'Abstract'),
        ('num_references', 'Number of references'),
        ('main_heading', 'Main heading'),
        ('index_keywords', 'Controlled terms'),
        ('author_keywords', 'Uncontrolled terms'),
        ('classification_code', 'Classification code'),
        ('doi', 'DOI'),
        ('funding', 'Funding details'),
        ('funding_text', 'Funding text'),
    ]

    for ukey, ei_label in field_map:
        val = unified.get(ukey, '')
        if val:
            lines.append(f'{ei_label}:{val}')

    lines.append('')
    lines.append('Database:Compendex')
    lines.append('Compilation and indexing terms, Copyright 2026 Elsevier Inc.')

    return '\n'.join(lines)


def export_merged_txt(
    unified_records: List[dict],
    template_path: str,
    output_path: str,
    template_db: str = '',
):
    encoding = detect_encoding(template_path)
    template_text = read_file_text(template_path, encoding)

    if template_db == 'wos' or 'Clarivate' in template_text[:200] or 'FN ' in template_text[:10]:
        converted = _convert_authors_for_export(unified_records, 'wos', 'txt')
        with open(output_path, 'w', encoding=encoding, errors='replace') as f:
            f.write('FN Clarivate Analytics Web of Science\nVR 1.0\n')
            for rec in converted:
                f.write(_build_wos_txt_record(rec))
                f.write('\n\n')
            f.write('EF')

    elif template_db == 'scopus' or 'Scopus' in template_text[:20]:
        from utils import detect_scopus_txt_language
        lang = detect_scopus_txt_language(template_text)
        header_lines = template_text.split('\n')[:2]
        header = '\n'.join(header_lines) + '\n\n'

        converted = _convert_authors_for_export(unified_records, 'scopus', 'txt')
        with open(output_path, 'w', encoding=encoding, errors='replace') as f:
            f.write(header)
            for rec in converted:
                f.write(_build_scopus_txt_record(rec, lang))
                f.write('\n\n')

    elif template_db == 'ei' or '<RECORD' in template_text[:200]:
        converted = _convert_authors_for_export(unified_records, 'ei', 'txt')
        with open(output_path, 'w', encoding=encoding, errors='replace') as f:
            for i, rec in enumerate(converted, start=1):
                f.write(_build_ei_txt_record(rec, i))
                f.write('\n\n')

    else:
        converted = _convert_authors_for_export(unified_records, 'wos', 'txt')
        with open(output_path, 'w', encoding=encoding, errors='replace') as f:
            for rec in converted:
                f.write(_build_wos_txt_record(rec))
                f.write('\n\n')


def export_merged(
    unified_records: List[dict],
    template_path: str,
    output_path: str,
    template_db: str = '',
):
    ext = os.path.splitext(template_path)[1].lower()
    if ext == '.csv':
        export_merged_csv(unified_records, template_path, output_path, template_db)
    elif ext in ('.xls', '.xlsx'):
        return export_merged_xls(unified_records, template_path, output_path, template_db)
    elif ext == '.txt':
        export_merged_txt(unified_records, template_path, output_path, template_db)
    return output_path
